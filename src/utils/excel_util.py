from typing import List, Any
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelUtil:
    """
    Excel 文件处理工具类。
    
    功能：
    - 创建 Excel 文件
    - 写入数据
    - 设置样式
    - 保存文件
    """
    
    @staticmethod
    def create_workbook() -> Workbook:
        """
        创建新的工作簿。
        
        返回：
            openpyxl.Workbook 实例
        """
        return Workbook()
    
    @staticmethod
    def create_sheet(workbook: Workbook, name: str) -> Worksheet:
        """
        创建工作表。
        
        参数：
            workbook: 工作簿实例
            name: 工作表名称
        
        返回：
            Worksheet 实例
        """
        if 'Sheet' in workbook.sheetnames:
            worksheet = workbook['Sheet']
            worksheet.title = name
        else:
            worksheet = workbook.create_sheet(title=name)
        return worksheet
    
    @staticmethod
    def write_data(sheet: Worksheet, data: List[List[Any]], start_row: int = 1, start_col: int = 1):
        """
        写入数据到工作表。
        
        参数：
            sheet: 工作表实例
            data: 二维数据列表
            start_row: 起始行号（从 1 开始）
            start_col: 起始列号（从 1 开始）
        """
        for row_idx, row_data in enumerate(data):
            for col_idx, cell_value in enumerate(row_data):
                sheet.cell(
                    row=start_row + row_idx,
                    column=start_col + col_idx,
                    value=cell_value
                )
    
    @staticmethod
    def write_headers(sheet: Worksheet, headers: List[str], row: int = 1):
        """
        写入表头。
        
        参数：
            sheet: 工作表实例
            headers: 表头列表
            row: 行号
        """
        for col_idx, header in enumerate(headers):
            sheet.cell(row=row, column=col_idx + 1, value=header)
    
    @staticmethod
    def set_column_width(sheet: Worksheet, column: str, width: int):
        """
        设置列宽。
        
        参数：
            sheet: 工作表实例
            column: 列标识（如 'A', 'B'）
            width: 宽度值
        """
        sheet.column_dimensions[column].width = width
    
    @staticmethod
    def auto_fit_columns(sheet: Worksheet):
        """
        自动调整所有列宽。
        
        参数：
            sheet: 工作表实例
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def apply_header_style(sheet: Worksheet, row: int = 1):
        """
        应用表头样式。
        
        参数：
            sheet: 工作表实例
            row: 表头行号
        """
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in sheet[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    @staticmethod
    def save(workbook: Workbook, path: str):
        """
        保存工作簿。
        
        参数：
            workbook: 工作簿实例
            path: 保存路径
        """
        workbook.save(path)
